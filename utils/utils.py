# coding: utf-8
# pylint: disable = invalid-name, W0105, C0111, C0301
"""Clustering by using scikit-learn libraries"""

from __future__ import absolute_import

# get relative data folder
import os,sys
ROOT_PATH = os.path.dirname(os.getcwd())
sys.path.insert(0,ROOT_PATH)
DATA_PATH = ROOT_PATH + '/data'

import os
from pathlib import Path
import traceback

from tqdm import tqdm
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D
import time

import sys
import numpy as np
import pandas as pd
from sklearn.model_selection import StratifiedShuffleSplit
from sklearn.neighbors import KernelDensity
from sklearn.cluster import MeanShift,estimate_bandwidth
import _pickle as pickle
from datetime import datetime
import glob
import re
from openpyxl import load_workbook



### Check if the script is run in Jupyter notebook
def isnotebook():
    
    try:
        get_ipython
        return True
    except:
        return False

    
    
class StopWatch(object):
    '''
    Stop Watch: keep track of time
    '''
    
    def __init__(self, reset=True):
        """
        Initialize the parameters.
        Load rules from a file
        """
        

        self.reset()
        
        pass
    
    
    def reset(self):
        '''
        Start the time tracking
        '''
        
        self.time_start = time.time()
        
        return self.time_start
    
    
    def check(self, reset=False):
        '''
        Check the time lapse
        '''
        if reset:
            time_start = self.time_start
            return (self.reset() - time_start)
        else:
            return (time.time() - self.time_start)


def excel_store(data,path=None,directory=None,module_name=None,file_name=None,prefix=None,sheet_name='default',add_sheet=False,append=False,verbose=False,**kwargs):
    """
    Store a pandas DataFrame to an excel file.
    
    Parameters
    ----------
    data : object
        data to be stored in pickle format.
        
    path : str
        absolute or relative (from home dir) path to store the pickle file
    
    directory : str
        absolute path to the directory in which to store the pickle file
        
    mudlue_name : str
        relative folder path under DATA_PATH, in which to store the picke file
        ignored if "directory" is specified
        
    file_name : str
        file name to store the pickle file
        
    prefix : str
        first part (=prefix) of the file name to store the pickle file
        ignored if "file_name" is specified
        
    sheet_name : str
        excel worksheet name
        
    add_sheet : bool
        if True, then keep the existing worksheet and add a new one
        
    append : bool
        if True, then append the data to the existing worksheet
        
    verbose : bool
        if True, then print progress
        
    Returns       
    -------
    result : boolean
        True if stored normally, False otherwise
        
    """
    if path: # the path is specified; use it.
        # convert relative path to absolute path
        path = add_homedir(path)
    else:    # the path is not specified
        # set the directory to store the file
        if directory:     data_dir = add_homedir(directory)
        elif module_name: data_dir = DATA_PATH + '/' + module_name
        else:             data_dir = DATA_PATH
        # set the file name to store the pickle file
        if file_name:     path     = data_dir + '/' + file_name
        elif prefix :        path  = data_dir + '/' + prefix + '_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.pkl'
        else:             path     = data_dir + '/dump_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.xlsx'

    # if the directory doesn't exist, create one
    if not os.path.exists(os.path.dirname(path)):
        try:
            os.makedirs(os.path.dirname(path))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise
    if append:
        # first read the file
        try:
            if verbose: print('Load excel file from {}'.format(path))
            with open(path, "rb") as fp:
                try:
                    data_old = pd.read_excel(fp,sheet_name=sheet_name,engine='openpyxl',verbose=False,**kwargs)
                    data = pd.concat([data_old,data])
                except:
                    data = data
        except IOError as e:
            print('File is new, resulting in I/O error({0}): {1}'.format(e.errno, e.strerror))
            pass
        except ValueError:
            print('Could not convert data to an integer.')
            pass
        except:
            print('Unexpected error:{}'.format(sys.exc_info()[0]))
            raise
        
    # excel store
    try:
        if verbose: print('Store excel file to {}'.format(path))
        # open a book if the file exists
        if add_sheet: 
            if os.path.isfile(path): 
                book = load_workbook(path)
                if append:
                    # we don't want to keep on adding the appended sheets.
                    # instead we want to delete the existing sheet, and add the appended sheet.
                    sheets = book.get_sheet_names()                # get the list of all the existing sheets
                    if sheet_name in sheets:                       # the sheet_name already exists
                        std = book.get_sheet_by_name(sheet_name)   # get the worksheet
                        book.remove_sheet(std)                     # remove it
            else                   : book = None
        if verbose: print('loaded workbook')
        # open an ExcelWriter for the file
        with pd.ExcelWriter(path,engine='openpyxl') as writer:  
            if verbose: print('oepned the file for writing')
            if add_sheet and book: writer.book = book
            if verbose: print('added worksheet')
            data.to_excel(writer,sheet_name=sheet_name,index=False)
        if verbose: print('wrote data to the file')
        return True
    except IOError as e:
        print('I/O error({0}): {1}'.format(e.errno, e.strerror))
        pass
    except ValueError:
        print('Could not convert data to an integer.')
        pass
    except:
        print('Unexpected error:{}'.format(sys.exc_info()[0]))
        raise
        
    return False
        

def excel_load(path=None,directory=None,module_name=None,file_name=None,prefix=None,sheet_name=None,append=False,verbose=False,**kwargs):
    """
    Load a pandas DataFrame from an excel file.
    
    Parameters
    ----------
    data : object
        data to be stored in pickle format.
        
    path : str
        absolute or relative (from home dir) path to store the pickle file
    
    directory : str
        absolute path to the directory in which to store the pickle file
        
    mudlue_name : str
        relative folder path under DATA_PATH, in which to store the picke file
        ignored if "directory" is specified
        
    file_name : str
        file name to store the pickle file
        
    prefix : str
        first part (=prefix) of the file name to store the pickle file
        ignored if "file_name" is specified
        
    verbose : bool
        if True, then print progress
        
    Returns       
    -------
    result : list,a pandas DataFrame
        If sheet_name is specified, then return the worksheet as a DataFrame.
        If sheet_name is None, then return a list of pandas Dataframes, 
        each of which corresponds to each of the worksheets in the excel file.
        
    """

    if path: # the path is specified; use it.
        # convert relative path to absolute path
        path = add_homedir(path)
    else:    # the path is not specified
        # set the directory to store the file
        if directory:     data_dir = add_homedir(directory)
        elif module_name: data_dir = DATA_PATH + '/' + module_name
        else:             data_dir = DATA_PATH
        # set the file name to store the pickle file
        if file_name:     path     = data_dir + '/' + file_name
        elif prefix :        path  = data_dir + '/' + prefix + '_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.pkl'
        else:             path     = data_dir + '/dump_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.xlsx'

    # if the directory doesn't exist, create one
    if not os.path.exists(os.path.dirname(path)):
        try:
            os.makedirs(os.path.dirname(path))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise

    # if the file doesn't exist, return False
    if not os.path.exists(path):
        return False
                
    # load the excel file
    try:
        if verbose: print('Load excel file from {}'.format(path))
        # load the workbook if sheet_name is not specified
        if not sheet_name: book = load_workbook(path)
        else             : book = None
        # open and read the excel sheet
        with open(path, "rb") as fp:
            if not book: work_sheets = pd.read_excel(fp,sheet_name=sheet_name,**kwargs)
            else:   # load all the worksheets
                work_sheets = {}
                for name in book.sheetnames: work_sheets[name] = pd.read_excel(fp,sheet_name=name,engine='openpyxl',**kwargs)
        return work_sheets
    except IOError as e:
        print('I/O error({0}): {1}'.format(e.errno, e.strerror))
        pass
    except ValueError:
        print('Could not convert data to an integer.')
        pass
    except:
        print('Unexpected error:{}'.format(sys.exc_info()[0]))
        raise
    
    return False
        

def pickle_store(data,path=None,directory=None,module_name=None,file_name=None,prefix=None,verbose=False):
    """
    Store a class object to a picle file.
    
    Parameters
    ----------
    data : object
        data to be stored in pickle format.
        
    path : str
        absolute or relative (from home dir) path to store the pickle file
    
    directory : str
        absolute path to the directory in which to store the pickle file
        
    mudlue_name : str
        relative folder path under DATA_PATH, in which to store the picke file
        ignored if "directory" is specified
        
    file_name : str
        file name to store the pickle file
        
    prefix : str
        first part (=prefix) of the file name to store the pickle file
        ignored if "file_name" is specified
        
    verbose : bool
        if True, then print progress
        
    Returns       
    -------
    result : boolean
        True if stored normally, False otherwise
        
    """

    if path: # the path is specified; use it.
        # convert relative path to absolute path
        path = add_homedir(path)
    else:    # the path is not specified
        # set the directory to store the file
        if directory:     data_dir = add_homedir(directory)
        elif module_name: data_dir = DATA_PATH + '/' + module_name
        else:             data_dir = DATA_PATH
        # set the file name to store the pickle file
        if file_name:        path  = data_dir + '/' + file_name
        elif prefix :        path  = data_dir + '/' + prefix + '_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.pkl'
        else        :        path  = data_dir + '/dump_' + datetime.now().strftime("%Y-%m-%d_%H:%M:%S") + '.pkl'

    if verbose: print('save pickle file to {}'.format(path))
    
    # if the directory doesn't exist, create one
    if not os.path.exists(os.path.dirname(path)):
        try:
            os.makedirs(os.path.dirname(path))
        except OSError as exc: # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise

    # pickle store
    try:
        with open(path, "wb") as fp:   #Pickling
            pickle.dump(data, fp)
        return True
    except IOError as e:
        print('I/O error({0}): {1}'.format(e.errno, e.strerror))
        pass
    except ValueError:
        print('Could not convert data to an integer.')
        pass
    except:
        print('Unexpected error:{}'.format(sys.exc_info()[0]))
        raise
        
    return False
        

def pickle_load(path=None,directory=None,module_name=None,file_name=None,prefix=None,verbose=False):
    """
    Load a class object to a picle file.
    
    Parameters
    ----------
    data : object
        data to be stored in pickle format.
        
    path : str
        absolute or relative (from home dir) path to load the pickle file from
    
    directory : str
        absolute path to the directory from which to load the pickle file
        
    mudlue_name : str
        relative folder path under DATA_PATH, from which to load the picke file
        ignored if "directory" is specified
        
    file_name : str
        file name to load the pickle file
        
    prefix : str
        first part (=prefix) of the file name to load the pickle file
        ignored if "file_name" is specified
        
        
    Returns       
    -------
    object : object
        data ot module to be restored from the storage
        
    """

    if path: # the path is specified; use it.
        # convert relative path to absolute path
        path = add_homedir(path)
    else:    # the path is not specified
        # set the directory to store the file
        if directory:     data_dir = add_homedir(directory)
        elif module_name: data_dir = DATA_PATH + '/' + module_name
        else:             data_dir = DATA_PATH
            
        if verbose: print("data_dir = {}".format(data_dir))
        
        # set the file name to store the pickle file
        if file_name:     path     = data_dir + '/' + file_name
        else: # search pickled report files
            pickled_report_files = []
            time_stamps = {}
            datetime_object = None
            if not prefix: prefix='dump'
            search = data_dir + '/' + prefix + '_*.pkl'
            for fn in glob.glob(search):
                pickled_report_files.append(fn)
                m = re.search(prefix+'_(.+?).pkl',fn)
                if m: 
                    datetime_object = datetime.strptime(m.group(1), '%Y-%m-%d_%H:%M:%S')
                    time_stamps[datetime_object] = fn
            # select the latest one
            if time_stamps: path = time_stamps[max(time_stamps.keys())]
            # otherwise return None
            else:
                print('Cannot find a pickle file in the folder {}.'.format(data_dir))
                return None

    if verbose: print('Load pickle file from {}'.format(path))

    # if the path exist, return False
    if not os.path.exists(path):
        print('path does not exist: {}'.format(path))
        return False
                

    # pickle store
    try:
        with open(path, "rb") as fp:   #Pickling
            return pickle.load(fp)
    except IOError as e:
        print('I/O error({0}): {1}'.format(e.errno, e.strerror))
        pass
    except ValueError:
        print('Could not convert data to an integer.')
        pass
    except:
        print('Unexpected error:{}'.format(sys.exc_info()[0]))
        raise
        
    return False
        

def exist_variable(x):
    """Check if the variable already exists"""
    
    try:
        _ = x.shape
        return True
    except NameError:
        print('Variable does not exist.')
        return False
    except AttributeError:
        print('Variable does not have a shape property.')    
        return False
    
    pass


def add_homedir(path):
    """
    If the path (string) starts with "~" then replace that with the absolute
    home path of the user.
    """
    
    # set the directory structure
    if path[0]=='~':
        path = str(Path.home()) + path[1:]
        
    return path



def select_stratified_samples(X,y,n_splits=1,N_max=None,random_state=0):
    '''
    Generate stratifeid data samples from the combination of feature and label data.

    parameters
    ----------
    X : feature data input
    y : label data input
    n_splits : int, default=1
        Number of re-shuffling & splitting iterations.
    N_max : int, default=None
        If float, should be between 0.0 and 1.0 and represent the proportion of 
        the dataset to include in the test split.
        If int, represents the absolute number of test samples. If None, the 
        value is set to the complement of the train size. If train_size is also 
        None, it will be set to 0.1.

    return
    ------
    indexes : list
        A list of numpy array, containing the IDs of stratified feature/label data samples.
    '''

    # geberate an iterator
    sss = StratifiedShuffleSplit(n_splits=n_splits, train_size=N_max, test_size=N_max, random_state=random_state)

    # Generate 
    indexes = []
    for train_index, test_index in sss.split(X, y):
        indexes.append(test_index)

    if n_splits==1: return indexes[0].tolist()
    else:           return indexes

    pass
    
    
def select_by_labels(X,y,labels):
    '''
    Select and compress the data samples and the labels by the label values.

    parameters
    ----------
    X : nd_array
        feature data input
    y : nd_array
        label data input
    labels : list
        Only use data samples whose labels are in this list.

    return
    ------
    X : compressed feature data input whose corresponding labels in y are in the list of labels.
    y : compressed label data input whose values are in the list of labels. 

    '''

    # reset the list of conditions and choices
    condlist   = []
    choicelist = []

    # set the list of conditions and choices
    for l in labels:
        condlist.append(y==l)
        choicelist.append(True)

    # Select the labels and create a condition array with True and False values.
    # If the label of y is in the list "labels" then the corresponding element in the "condition" array is True.
    condition = np.select(condlist,choicelist,default=False).reshape(y.shape[0])

    # Compress the array according to the condition element.
    X_compressed = np.compress(condition,X,axis=0)
    y_compressed = np.compress(condition,y,axis=0)

    return X_compressed, y_compressed
    

def correct_kwargs(kwargs):
    '''
    scan kwargs and correct known characters
    
    parameters
    ----------
    kwargs : dict
        input arguments

    return
    ------
    kwargs : dict
        corrected arguments

    '''
    
    for k in kwargs.keys():
        if kwargs[k]=='None':  kwargs[k] = None
        if kwargs[k]=='True':  kwargs[k] = True
        if kwargs[k]=='False': kwargs[k] = False

    return kwargs


def print_formated(variable,value=None,value_format='',value_unit=''):
    """print variable values in a controlled format

    Parameters
    ----------
    variable: string with length <= 30
    value: string, float, int, etc.
    value_format: format for the value

    Returns
    -------

    """

    _format = '{:30}{:^3}{' + value_format + '} {}'
    print(_format.format(variable,'=',value,value_unit))

    pass


def print_debug(frameinfo,variable,value=None,value_format='',value_unit='',func=False,new_line=True,err=None):
    """print variable values in a controlled format

    Parameters
    ----------
    variable: string with length <= 30
    frameinfo : frameinfo object
    value: string, float, int, etc.
    value_format: format for the value
    value_unit: string denoting the unit
    new_line: if True, change line when showing the value
    err : if exists, show the traceback

    Returns
    -------

    """
    
    if err: traceback.print_tb(err.__traceback__)

    lf = '\n' if new_line else ''
    fn = '{:<30}' if func else ''
    _format = lf+'file:{}, line:{}, '+fn+'{} = '+lf+'{' + value_format + '} {}'
    
    if func: print(_format.format(frameinfo.filename,frameinfo.lineno,frameinfo.function,variable,value,value_unit))
    else   : print(_format.format(frameinfo.filename,frameinfo.lineno,variable,value,value_unit))

    pass


def get_value_from_tensor_rv(rv,is_tuple=False):
        '''
            Get value (or numpy.array) from PyMC3 FreeRV or Theano tensors.

            parameters
            ----------
            rv : Theano tensor ,PyMC3 FreeRV, or tuple
                random variable in tensor format
                in case is_tuple=True, then this is a list of random variables
                
            is_tuple : bool
                if True, rv is a list of random variables.
                default is False
                
            returns
            -------
            val : int, float or numpy.array
                random variable value

        '''
        
        if is_tuple or ("tuple" in str(type(rv))):
            val = []
            for rv_ in rv: val.append(get_value_from_tensor_rv(rv_,is_tuple=False))
            val = tuple(val)
        elif ("list" in str(type(rv))):
            val = []
            for rv_ in rv: val.append(get_value_from_tensor_rv(rv_,is_tuple=False))
            val = tuple(val)
        else:
            str_type = str(type(rv))
            if ('pymc3.model.FreeRV' in str_type) or ('theano.tensor' in str_type):
                try:    val = rv.eval()
                except: val = rv.tag.test_value
            else:
                try:    val = rv
                except: sys.exit('type(rv)={}: unknown type'.format(str_type))

            if 'numpy.ndarray' in str(type(val)):
                if val.ndim==0: val = float(val)

                
        return val

    
def debug_print_FreeRV(debug_cond,debug_level,frameinfo,variable,value=None,value_format='',value_unit='',
                       func=False,new_line=True,err=None,print_type=True,print_dtype=True,print_shape=True,print_value=True,
                       debug_level_value=None,debug_level_type=None,debug_level_dtype=None,debug_level_shape=None):
    '''
        Print the value of theano.tensor or PyMC3 FreeRV, along with debug info

        parameters
        ----------
        debug_cond : int
            maximum debug_level value
            if debug_level<=debug_cond, then print out the debug info
        debug_level : int
            value which is compared against debug_cond to decide whether or not to print
        frameinfo : dict
            containing debug info
        variable : str
            variable name
        value : int, float, list, numpy.array, etc.
            value to be printed
        value_format : str
            print format
        value_unit : str
            unit of the value, to print
        func : bool
            if True, then print the function name from which this "debug_print_FreeRV" function is called
        new_line : bool
            if True, then change line when printing the value
        err : bool
            if True, then print error information
        print_shape : bool
            if True, then print the shape of the variable
        print_value : bool
            if True, then print the value
        print_type : bool
            if True, then print the type of the variable
        debug_level_value : int
            value which is compared against debug_cond to decide whether or not to print the value
            if None, then it is the same as debug_level
        debug_level_type : int
            value which is compared against debug_cond to decide whether or not to print the type
            if None, then it is the same as debug_level
        debug_level_dtype : int
            value which is compared against debug_cond to decide whether or not to print the dtype
            if None, then it is the same as debug_level
        debug_level_shape : int
            value which is compared against debug_cond to decide whether or not to print the shape
            if None, then it is the same as debug_level

        returns
        -------

    '''
    if not debug_level_value: debug_level_value = debug_level
    if not debug_level_type : debug_level_type  = debug_level
    if not debug_level_dtype: debug_level_dtype = debug_level
    if not debug_level_shape: debug_level_shape = debug_level
    
    str_type_org = str(type(value))
    val = get_value_from_tensor_rv(value)
    if print_type and (debug_level_type<=debug_cond):
        print_debug(frameinfo,variable='type('+variable+')',value=str_type_org,value_format=value_format,func=func,new_line=False,err=err)
    if print_dtype and (debug_level_dtype<=debug_cond):
        try:
            print_debug(frameinfo,variable=variable+'.dtype',value=str(value.dtype),value_format=value_format,func=func,new_line=False,err=err)
        except:
            print_debug(frameinfo,variable='dtype not available',value=str(type(value)),value_format=value_format,func=func,new_line=False,err=err)
    if 'numpy.ndarray' in str_type_org:
        if print_shape and (debug_level_shape<=debug_cond):
            print_debug(frameinfo,variable=variable+'.shape',value=value.shape,value_format=value_format,func=func,new_line=False,err=err)
    else: # assume Theano tensor
        try:
            if print_shape and (debug_level_shape<=debug_cond):
                print_debug(frameinfo,variable=variable+'.shape',value=val.shape,value_format=value_format,func=func,new_line=False,err=err)
        except:
            print_debug(frameinfo,variable='cannot print shape for the type',value=str(type(val)),value_format=value_format,func=func,new_line=False,err=err)
    if print_value and (debug_level_value<=debug_cond): 
        print_debug(frameinfo,variable=variable,value=val,value_format=value_format,func=func,new_line=new_line,err=err)


def get_value_from_tensor(var,is_tuple=False):
        '''
            Get value (or numpy.array) from PyMC3 FreeRV or Theano tensors.

            parameters
            ----------
            val : Tensorflow tensor, or tuple
                random variable in tensor format
                in case is_tuple=True, then this is a list of random variables
                
            is_tuple : bool
                if True, var is a list of random variables.
                default is False
                
            returns
            -------
            val : int, float or numpy.array
                random variable value

        '''
        
        if is_tuple or ("tuple" in str(type(var))):
            val = []
            for var_ in var: val.append(get_value_from_tensor(var_,is_tuple=False))
            val = tuple(val)
        elif ("list" in str(type(var))):
            val = []
            for var_ in var: val.append(get_value_from_tensor(var_,is_tuple=False))
        else:
            str_type = str(type(var))
            if ('pymc3.model.FreeRV' in str_type) or ('theano.tensor' in str_type):
                try:    val = var.eval()
                except: val = var.tag.test_value
            else:
                try:    val = var
                except: sys.exit('type(var)={}: unknown type'.format(str_type))

            if 'numpy.ndarray' in str(type(val)):
                if val.ndim==0: val = float(val)

                
        return val


def debug_print_tensor(debug_cond,debug_level,frameinfo,variable,value=None,value_format='',value_unit='',
                       func=False,new_line=True,err=None,print_type=True,print_dtype=True,print_shape=True,print_value=True,
                       debug_level_value=None,debug_level_type=None,debug_level_dtype=None,debug_level_shape=None):
    '''
        Print the value of theano.tensor or PyMC3 FreeRV, along with debug info

        parameters
        ----------
        debug_cond : int
            maximum debug_level value
            if debug_level<=debug_cond, then print out the debug info
        debug_level : int
            value which is compared against debug_cond to decide whether or not to print
        frameinfo : dict
            containing debug info
        variable : str
            variable name
        value : int, float, list, numpy.array, etc.
            value to be printed
        value_format : str
            print format
        value_unit : str
            unit of the value, to print
        func : bool
            if True, then print the function name from which this "debug_print_FreeRV" function is called
        new_line : bool
            if True, then change line when printing the value
        err : bool
            if True, then print error information
        print_shape : bool
            if True, then print the shape of the variable
        print_value : bool
            if True, then print the value
        print_type : bool
            if True, then print the type of the variable
        debug_level_value : int
            value which is compared against debug_cond to decide whether or not to print the value
            if None, then it is the same as debug_level
        debug_level_type : int
            value which is compared against debug_cond to decide whether or not to print the type
            if None, then it is the same as debug_level
        debug_level_dtype : int
            value which is compared against debug_cond to decide whether or not to print the dtype
            if None, then it is the same as debug_level
        debug_level_shape : int
            value which is compared against debug_cond to decide whether or not to print the shape
            if None, then it is the same as debug_level

        returns
        -------

    '''
    if not debug_level_value: debug_level_value = debug_level
    if not debug_level_type : debug_level_type  = debug_level
    if not debug_level_dtype: debug_level_dtype = debug_level
    if not debug_level_shape: debug_level_shape = debug_level
    
    str_type_org = str(type(value))
    val = get_value_from_tensor(value)
    if print_type and (debug_level_type<=debug_cond):
        print_debug(frameinfo,variable='type('+variable+')',value=str_type_org,value_format=value_format,func=func,new_line=False,err=err)
    if print_dtype and (debug_level_dtype<=debug_cond):
        try:
            print_debug(frameinfo,variable=variable+'.dtype',value=str(value.dtype),value_format=value_format,func=func,new_line=False,err=err)
        except:
            print_debug(frameinfo,variable='dtype not available',value=str(type(value)),value_format=value_format,func=func,new_line=False,err=err)
    if 'numpy.ndarray' in str_type_org:
        if print_shape and (debug_level_shape<=debug_cond):
            print_debug(frameinfo,variable=variable+'.shape',value=value.shape,value_format=value_format,func=func,new_line=False,err=err)
    else: # assume Theano tensor
        try:
            if print_shape and (debug_level_shape<=debug_cond):
                print_debug(frameinfo,variable=variable+'.shape',value=val.shape,value_format=value_format,func=func,new_line=False,err=err)
        except:
            print_debug(frameinfo,variable='cannot print shape for the type',value=str(type(val)),value_format=value_format,func=func,new_line=False,err=err)
    if print_value and (debug_level_value<=debug_cond): 
        print_debug(frameinfo,variable=variable,value=val,value_format=value_format,func=func,new_line=new_line,err=err)


